from typing import Any, List
from typing import Literal
import librosa
import matplotlib.axes
from scipy import signal
import numpy as np
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import scipy.spatial
import os
import seaborn as sns
import time

color = {
    'blue': (0, 0.4470, 0.7410),
    'orange': (0.8500, 0.3250, 0.0980),
    'red': (0.6350, 0.0780, 0.1840),
    'violet': (0.4940, 0.1840, 0.5560),
    'green': (0.4660, 0.6740, 0.1880),
    'cyan': (0.3010, 0.7450, 0.9330)
}

def largestpowerof2(n:int):
    '''
    return a number which is the largest power of 2 and less than n
    '''
    while(n & (n - 1)):
        n &= (n - 1)
    return n

def butter_highpass(input, t, cutoff, fs, order = 5, axis = 0, visualize = False):
    sos = signal.butter(order, cutoff, btype='highpass', fs=fs, output='sos')
    output = signal.sosfilt(sos, input, axis=axis)
    if visualize:
        fig, (ax1, ax2) = plt.subplots(2, 1, sharex=True, layout="tight")
        ax1.plot(t,input)
        ax1.set_title('input signal')
        ax2.plot(t, output)
        ax2.set_title('After %d hz high-pass filter'%cutoff)
        ax2.set_xlabel('Time [seconds]')
        plt.show()
    return output

def fft(df:pd.DataFrame, fs = 1, nperseg=8192, window = np.hanning(8192), noverlap=8192//2, fg_column=3, pulse_per_round=2, 
        nfft=None, domain: Literal["frequency", "order"] = "frequency", cols=3, **arg):
    """
    do FFT with window frame, return complex number of each window frame
    
    parameters
    -------
    df : input acc data, each column represents a sequence signal of accelerometers, and the last column is the fg sensor.
    fs : sampling frequency
    nperseg : number of samples of each window frame. In the order domain, this is estimated by fs/average_rotating_frequency.
    noverlap : overlapping samples. In the order domain, this is used for time synchronized average for every noverlap cycles.
    fg_column : fg signal as square wave, usually the last column number of input data frame.
    pulse_per_round : fg sensor pulse numbers per round
    nfft : int, optional
        Length of the FFT used, if a zero padded FFT is desired. If
        `None`, the FFT length is `nperseg`. Defaults to `None`.
        controlled the frequency resolution of the spectrum, using zero-padding to increase the resolution as nfft/nperseg
    domain : units of the spectrum x labels
        * 'frequency': hz
        * 'order': transfer the angular rotation to order. The horizontal coordinate of the output spectrum is the amplitude 
                   versus the multiples of the inner race rotation cycle, fr. 
    cols : use first int(cols) for fft analysis, this values usually equal to the number of accelerometers.
    
    Note
    ------
    1 . If we divide frequency with rotation frequency to get order, the corresponding order is varying from each frame, that cause the
        difficulties for the following averaging step. we can't simply get the mean of the transformed result of each frame.
    2 . (Crucial )use the fourier transformed result to do cross spectrum, the complex number with varying phase caused by the equal 
        frame length will cancel each other and cause the csd result vary small and shows no relation to order.
    
    Reference
    -------
    **order domain analysis**
    Transform the equally time stepped samples into equally angular stepped samples for analysis rotating vibration
    (https://dsp.stackexchange.com/questions/42345/time-synchronous-averaging-matlab)
    """
    # seperate the original signal into frames
    if domain == 'frequency':
        frames = librosa.util.frame(df, frame_length=nperseg, hop_length=int(nperseg-noverlap), axis=0)
    elif domain == 'order':
        frames = slice_frame(df, fg_column=fg_column, threshold=0, pulse_per_round=pulse_per_round, cols=cols, **arg)
        # detrend again for the time synchronous averaged frame
        signal.detrend(frames, axis=1, type='constant', overwrite_data=True)
        #fs = nperseg if nfft == None else nfft

    # reshape if the frame length is unequal to nfft 
    if frames.shape[1] != nfft:
        # get the reshaped frame_numbers
        scale = nfft / frames.shape[1]
        frame_number = int(frames.shape[0] - (frames.shape[0] % scale))
        frames = frames[:frame_number, :, :]
        frames = frames.reshape((int(frames.shape[0]/scale), nfft, -1))

    windowed_frames = np.empty(frames.shape) # frame.shape = [frame_numbers, frame_length, columns_dataframe]
    for col in range(frames.shape[-1]):
        np.multiply(window, frames[:,:,col], out=windowed_frames[:,:,col])
    sp = np.fft.rfft(windowed_frames, n=nfft, axis=1, norm='backward')
    freq = np.fft.rfftfreq(n=windowed_frames.shape[1] if nfft == None else nfft, d=1./fs)
    return freq, sp
    
def get_fft(df: pd.DataFrame, cut_off_freq = 0, fs = 48000, nperseg=8192, noverlap = 8192*0.75,
            domain: Literal["frequency", "order"] = "frequency", fg_column=3, pulse_per_round = 2, nfft = None, cols = None,
            average: Literal["rms", "mean", "median"] = "rms"):
    '''
    return fft as dataframe type
    
    parameters
    ------
    cut_off_freq : cut off frequency for high-pass filter
    average : the averaging method of windowed frames
    df : input acc data, each column represents a sequence signal of accelerometers, and the last column is the fg sensor.
    fs : sampling frequency
    nperseg : number of samples of each window frame. In the order domain, this is estimated by fs/average_rotating_frequency.
    noverlap : overlapping samples. In the order domain, this is used for time synchronized average for every noverlap cycles.
    fg_column : fg signal as square wave, usually the last column number of input data frame.
    pulse_per_round : fg sensor pulse numbers per round
    nfft : int, optional
        Length of the FFT used, if a zero padded FFT is desired. If
        `None`, the FFT length is `nperseg`. Defaults to `None`.
        controlled the frequency resolution of the spectrum, using zero-padding to increase the resolution as nfft/nperseg
    domain : units of the spectrum x labels
        * 'frequency': hz
        * 'order': transfer the angular rotation to order. The horizontal coordinate of the output spectrum is the amplitude 
                   versus the multiples of the inner race rotation cycle, fr. 
    cols : use first int(cols) for fft analysis, this values usually equal to the number of accelerometers.

    Note
    ------
    **signal processing step**
    1. subtract dc bias from accelerometer data
    2. high pass filter (optional)
    3. do FFT with hanning window frame
    4. get average of FFT spectrum based on different return type of fft helper function
    '''
    # subtract dc bias from acc data
    detrend_df = df - np.mean(df.to_numpy(), axis=0)
    if cut_off_freq > 0:
        # use high-pass filter to remove dc
        detrend_df = butter_highpass(detrend_df, df.index, cut_off_freq, fs, 2, visualize=True)
    freq, sp = fft(detrend_df, fs=fs, nperseg=nperseg, window=np.hanning(nperseg), noverlap=noverlap, fg_column=fg_column, pulse_per_round=pulse_per_round, 
                   nfft=nfft, domain=domain, cols=cols)
    
    # Average over windows
    sp = sp[:,:,:cols] # get the used range
    if average == 'rms':
        sp_averaged = np.sqrt(np.mean(np.power(np.abs(sp),2), axis=0))
    elif average == 'median':
        # np.median must be passed real arrays for the desired result
        if np.iscomplexobj(sp):
            sp_averaged = (np.median(np.real(sp), axis=0) + 1j * np.median(np.imag(sp), axis=0))
        else:
            sp_averaged = np.median(sp, axis=0)
    elif average == 'mean':
        sp_averaged = sp.mean(axis=0)
    else:
        raise ValueError('choose from specified methods')
    sp_averaged = pd.DataFrame(data=sp_averaged, columns=df.columns[:cols])
    if domain == 'frequency':
        sp_averaged['Frequency (Hz)'] = freq
        sp_averaged.set_index('Frequency (Hz)', inplace=True)
    elif domain == 'order':
        sp_averaged['order (cycle)'] = freq
        sp_averaged.set_index('order (cycle)', inplace=True)
            
    return sp_averaged

def annotatePeaks(x: Any, y: Any, ax: matplotlib.axes.Axes = None, prominence = None, dot = None, 
                  annotateX=True, annotateY=False,
                  rotation = 45, 
                  xytext=(0, 30), textcoords='offset pixels',
                  arrowprops = dict(facecolor='blue', arrowstyle="->", connectionstyle="arc3")):    
    """
    analysis the peak and add annotation on the graph

    parameters
    --------
    x : 1d array index
    y : 1d array (note: if it is a spectrum, remember to use absolute value)
    
    """
    peaks, dic = signal.find_peaks(y, prominence=prominence)
    if ax != None:
        for idx in peaks:
            if dot != None:
                ax.plot(x[idx], y[idx], dot)
            string = ''
            if annotateX:
                string += '%.2f'%x[idx]
            if annotateY:
                string += ', %.2f'%y[idx]
            ax.annotate(string, 
                        xy=(x[idx], y[idx]), rotation=rotation, xycoords='data',
                        xytext=xytext, textcoords=textcoords,
                        arrowprops=arrowprops)
    return peaks

def calc_rms(df: pd.DataFrame):
    rms = df.copy()**2
    rms = rms.mean()**0.5
    return rms

def stat_calc(df: pd.DataFrame):
    """
    calculate accelerate data peak, rms, crest factor and standard deviation
    """
    df_stats = pd.concat([df.abs().max(),calc_rms(df)],axis=1)
    df_stats.columns = ['Acceleration Peak (g)','Acceleration RMS (g)']
    df_stats['Crest Factor'] = df_stats['Acceleration Peak (g)'] / df_stats['Acceleration RMS (g)']
    df_stats['Standard Deviation (g)'] = df.std()
    df_stats.index.name = 'Data Set'
    return df_stats

def get_psd(df: pd.DataFrame, cut_off_freq: float = 0.0, fs: int = 1, nperseg: int = 8192, noverlap: int = 8192 * 0.75, 
            domain:Literal['frequency', 'order'] = 'frequency', nfft: int = None, cols: int = 3, 
            average:Literal['mean', 'median', 'None'] = 'mean',
            **arg):
    detrend_df = df - np.mean(df.to_numpy(), axis=0)
    index_name = ''
    if cut_off_freq > 0:
        detrend_df = butter_highpass(detrend_df, df.index, cut_off_freq, fs, 2)
    if domain == 'frequency':
        f, psd = signal.welch(detrend_df, fs=fs, nperseg=nperseg, window='hann', noverlap=noverlap, axis=0)
        index_name = 'Frequency (Hz)'
    elif domain == 'order':
        f, psd = csd_order(x=detrend_df, y=detrend_df, fs = fs, nperseg=nperseg, noverlap=noverlap, cols=cols, nfft=nfft, average=average, **arg)
        psd = np.real(psd)
        index_name = 'Order'
    if average == 'None':
        # use dictionary of dataframe
        df_psd = {}
        for sensor_channel in range(psd.shape[-1]):
            df_psd[df.columns[sensor_channel]] = pd.DataFrame(psd[:,:,sensor_channel], columns=f)
    else:
        df_psd = pd.DataFrame(psd,columns=df.columns[:cols])
        df_psd[index_name] = f
        df_psd = df_psd.set_index(index_name)
    return df_psd

def save_bar_plot(name: Any, value:Any, plot_title:str, file_name:str, figsize:tuple = (10, 10), path_dir:str = './fig/'):
    plt.style.use('fivethirtyeight')
    fig, ax = plt.subplots(layout="constrained", figsize=figsize)
    # Horizontal Bar Plot
    ax.barh(name, value)
    # Show top values 
    ax.invert_yaxis()
    # Add annotation to bars
    for i in range(len(value)):
        plt.text(value.iloc[i], i, 
                 str(round(value.iloc[i], 3)),
                 fontsize = 10, fontweight ='bold',verticalalignment="center",
                 color ='grey')
    
    # Add Plot Title
    ax.set_title(plot_title, loc ='left', fontsize=14)
    # save figure
    if not os.path.exists(path_dir):
        print('%s not exist, create new directory.'%path_dir)
        os.makedirs(path_dir)
    fig.savefig(path_dir+file_name, transparent=False, dpi=80, bbox_inches="tight")

def to_excel(df:pd.DataFrame, sheet_name:str, filename:str):
    '''
    use pandas to_excel(), if the file does not exist, an excel file with one default sheet is created.
    after saving, remove the first default blank sheet.
    ''' 
    # if file not exist
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        wb.save(filename)
        wb.close()
    with pd.ExcelWriter(filename, mode="a", if_sheet_exists="new", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name)
        if 'Sheet' in writer.book.sheetnames:    
            writer.book.remove(writer.book['Sheet'])
        writer.book.save(filename)
        writer.book.close()

def to_parquet(df:pd.DataFrame, suffix:str, filename:str):
    '''
    use same parameter structure as to_excel(), since parquet file format does not have sheet, use suffix instead.
    filename might contain .parquet using the same pattern as to_excel
    '''
    df = cast_column_to_str(df, ndigits=1) # fix ValueError: parquet must have string column names
    filename = filename.removesuffix('.gzip').removesuffix('.parquet')
    df.to_parquet(path=filename + '_' + suffix + '.parquet.gzip', compression='gzip')

def acc_processing_df(
        df: pd.DataFrame,
        analysis_mask: int,
        sheet_name: str,
        df_stats: pd.DataFrame, 
        coherence_compare_df: pd.DataFrame = None,
        fft_result_filename:str = 'fft.xlsx',
        psd_result_filename:str = 'psd.xlsx',
        Cxy_result_filename:str = 'coherence.xlsx',
        file_export_func = to_excel,
        **arg
        ):
    '''
    given an input as dataframe type of level vs time data, specificly accelerometer sensor data, do the selected analysis
    and output to specified xlsx file, each output is saved in one sheet of a workbook.

    Parameters
    ----------
    domain : all analysis use the same domain 
    '''
    export_funcs = [to_excel, to_parquet]

    if file_export_func not in export_funcs:
        raise ValueError("Unknown export option '{}', must be one of: {}"
                         .format(file_export_func, export_funcs))

    if analysis_mask & 0b0001:
        # state
        df_stats = pd.concat(df_stats, stat_calc(df))
    if analysis_mask & 0b0010:
        # fft
        df_fft = get_fft(df, **arg)
        file_export_func(df_fft, sheet_name, fft_result_filename)
    if analysis_mask & 0b0100:
        # psd
        df_psd = get_psd(df, **arg)
        if type(df_psd) is dict:
            for key, value in df_psd.items():
                file_export_func(value, key, psd_result_filename)
        else:
            file_export_func(df_psd, sheet_name, psd_result_filename)
    if analysis_mask & 0b1000:
        # coherence
        df_Cxy = coherence(x=coherence_compare_df, y=df, **arg)    
        file_export_func(df_Cxy, sheet_name, Cxy_result_filename)
        
def acc_processing_hdf(
        hdf_level_time_filename:str,
        analysis_mask: int, 
        comparing_sample_lr: str = '',
        comparing_sample_ud: str = '',
        state_result_filename:str = 'state.xlsx',
        fft_result_filename:str = 'fft.xlsx',
        psd_result_filename:str = 'psd.xlsx',
        Cxy_result_filename:str = 'coherence.xlsx',
        rename_column_method = None,
        usecols:list = None,
        sheets:list = None,
        **arg):
    """
    read level vs time acoustic .excel file, loop for each sheet or each file in the directory as panda data frame 
    and do selected processing, save the result into excel workbook.
    
    Parameters
    ----------
    hdf_level_time_filename : head acoustic exported excel file
    analysis_mask : whether use the data frame to calculate 
        * time domain standard deviation etc..
        * fft
        * psd
        * coherence
    
    comparing_sample_lr : str,
        comparing file for coherence analysis, when the sensors are attached at left, right, and axial side of the sample.
    comparing_sample_ud : str,
        comparing file for coherence analysis, when the sensors are attached at up, down, and axial side of the sample.
    """
    if analysis_mask & 0b0001:
        df_stats = pd.DataFrame()
    
    if analysis_mask & 0b1000:
        comparing_sample_lr_df = pd.read_excel(hdf_level_time_filename, header=0, index_col=0, skiprows=13, usecols=usecols, sheet_name=comparing_sample_lr)
        comparing_sample_ud_df = pd.read_excel(hdf_level_time_filename, header=0, index_col=0, skiprows=13, usecols=usecols, sheet_name=comparing_sample_ud)
    else:
        comparing_sample_lr_df = pd.DataFrame()
        comparing_sample_ud_df = pd.DataFrame()
        
    workbook = openpyxl.load_workbook(hdf_level_time_filename, read_only=True, data_only=True, keep_links=False)
    print("There are %d"%len(workbook.sheetnames) + " sheets in this workbook ( " + hdf_level_time_filename + " )")
    
    sheets = workbook.sheetnames if sheets == None else sheets
    for sheet in sheets:
        title = workbook[sheet]["B5"].value
        df = pd.read_excel(hdf_level_time_filename, sheet_name=sheet, header=0, index_col=0, skiprows=13, usecols=usecols)
        # rewrite column title adding title
        if rename_column_method is not None:
            rename_column_method(df, title)
        acc_processing_df(df = df, analysis_mask=analysis_mask, sheet_name=title,
                          df_stats=df_stats if analysis_mask & 0b0001 else None, 
                          coherence_compare_df=comparing_sample_lr_df if 'lr' in title else comparing_sample_ud_df,
                          fft_result_filename=fft_result_filename, psd_result_filename=psd_result_filename,
                          Cxy_result_filename=Cxy_result_filename, **arg)

    workbook.close()
    if analysis_mask & 0b0001:
        df_stats.to_excel(state_result_filename, sheet_name='state')

def compare_peak_from_fftdataframe(df: pd.DataFrame):
    peak_dict = {}
    for col_name in df.columns:
        series = df[col_name].to_numpy()
        peak_idxs = annotatePeaks(a=series, freq=df.index, prominence=0.25*series)
        update_peak_dic(peak_dict, peak_idxs)
    print("peak numbers: %d"%len(peak_dict.keys()))
    return peak_dict

def update_peak_dic(dic:dict, idxs:list[int]):
    for key in idxs:
        if key in dic.keys():
            dic[key] += 1
        else:
            dic[key] = 1

def rename_col(df: pd.DataFrame, title:str):
    # rewrite column title adding title
    df.rename(columns=lambda x: title[15:22] + '_' + x.split()[0][4:], inplace=True)

def read_sheets(filename:str, file_type:Literal['hdf', 'normal'] = 'normal', rename_column_method = None, usecols = None, combine = True,
                axis: int = 1, df:pd.DataFrame = None):
    """
    read previous exported excel file, loop for each sheet, combine as one pandas data frame, or return a dictionary of dataframe
    
    Parameters
    ----------
    file_type : 
        * 'hdf': head acoustic file format, which has a header before the data
        * 'normal': only the first row be the header
    rename_column_method : a rename function
    usecols : parameter for pd.read_excel
    combine : if True, combine the workbook sheets to a single dataframe
    axis : the pd.concat axis
    df : the existing df for combine
    """
    workbook = openpyxl.load_workbook(filename, read_only=True, data_only=True, keep_links=False)
    print("There are %d"%len(workbook.sheetnames) + " sheets in this workbook ( " + filename + " )")

    if file_type == 'normal':
        df_dict = pd.read_excel(filename, sheet_name=None, header=0, index_col=0, usecols=usecols)
    if file_type == 'hdf':
        df_dict = pd.read_excel(filename, sheet_name=None, header=0, index_col=0, skiprows=13, usecols=usecols)
        for sheet in workbook.sheetnames:   
            title = workbook[sheet]["B5"].value
            rename_column_method(df_dict[sheet], title)
    if combine == False:
        return df_dict 
    # combine all fft to the same dataframe
    df_all_fft = pd.DataFrame() if df is None else df
    for sheet in workbook.sheetnames:
        if axis == 0:
            # when combining sheets, add a column for its sheet_name
            df_dict[sheet].reset_index(drop=True, inplace=True)
            df_dict[sheet]['name'] = sheet
            df_all_fft = pd.concat([df_all_fft, df_dict[sheet]], axis=0)
        else:
            df_all_fft = pd.concat([df_all_fft, df_dict[sheet]], axis=1)
    workbook.close()
    return df_all_fft

def parse_digital(filename:str):
    '''
    parse 6-number digital sample number from filename
    '''
    for s in filename.split('_'):
        if s.isdigit():
            return s

def cast_column_to_str(df:pd.DataFrame, ndigits:int, labels:list = None):
    '''
    cast column labels to string. `Autosklearn` only accepts column label type int or string, 
    so we want to cast float into string to prevent fitting errors.
    when read dataframe from excel or parquet file, the float number is not precise, it need to be 
    rounded to a specified number of decimal places
    '''
    # make a dictionary for mapper
    mapper = {}
    
    if labels is None:
        for col in df.columns:
            if type(col) is not str:
                mapper[col] = str(round(col, ndigits))
    else:
        if len(df.columns) != len(labels):
            raise ValueError('df columns length does not match labels')
        for col, label in zip(df.columns, labels):
            mapper[col] = label
    #print(mapper)
    return df.rename(columns=mapper, copy=False)

def read_parquet_keyword(keyword: str, dir:str, parse_func:None):
    '''
    read keyword filtered parquet files, and combined as a whole dataframe
    
    Parameters
    ----------
    ketword : the selected channel as use case
    dir : the location of all parquet files
    parse_func : the function for getting the value of ['name'] column

    Examples
    --------
    >>>     df = read_parquet_keyword('ud_up', dir, parse_digital)
    '''
    df_all = pd.DataFrame()
    for file_name in os.listdir(dir):
        if '.parquet' in file_name and keyword in file_name:
            df = pd.read_parquet(dir+file_name)
            print("read %s"%file_name)
            if parse_func is not None:
                df['sample_num'] = parse_func(file_name)
            df_all = pd.concat([df_all, df], axis=0)
    df_all.reset_index(drop=True, inplace=True)
    return df_all

def class_average_peak(peak_dic:dict, df_fft: pd.DataFrame):
    idx_list = []
    for peak in peak_dic:
        freq = df_fft.index[peak]
        if freq > 5000:
            continue
        if peak_dic.get(peak) < 2:
            continue
        idx_list.append(peak)

    return df_fft.iloc[idx_list].mean()

def read_acc_file(path:str, usecols:list = None):
    '''
    read acc file of different file extension, currently only include .xlsx or .parquet
    '''
    if path.endswith('.parquet.gzip') or path.endswith('.parquet'):
        return pd.read_parquet(path = path, columns=usecols)
    if path.endswith('.xlsx'):
        return pd.read_excel(io=path, header=0, usecols=usecols)

def acc_processing_raw(
        dir:str,
        analysis_mask: int,
        comparing_sample_lr: str = '',
        comparing_sample_ud: str = '',
        state_result_filename:str = 'state.xlsx',
        fft_result_filename:str = 'fft.xlsx',
        psd_result_filename:str = 'psd.xlsx',
        Cxy_result_filename:str = 'coherence.xlsx',
        usecols:list = None,
        **arg):
    """
    read level vs time .xlsx or .parquet file, loop for each file in the directory, read as panda data frame and do selected processing, 
    save the result of selected processing to selected output format, if excel is chosen, the result will save into seperate excel sheet.
    
    Parameters
    ----------
    dir : directory where acc files are
    analysis_mask : whether use the data frame to calculate 
        * time domain standard deviation etc..
        * fft
        * psd
        * coherence
    
    comparing_sample_lr : str,
        comparing acc file for coherence analysis, when the sensors are attached at left, right, and axial side of the sample.
    comparing_sample_ud : str,
        comparing acc file for coherence analysis, when the sensors are attached at up, down, and axial side of the sample. 
    
    Examples
    --------
    >>>    acc_processing_raw(dir='../../test_data//20240911_good_samples//acc_data//', analysis_mask=0b1000, fs=1024, nperseg=int(51200/44), noverlap=10,
                         domain='order', nfft=1024, cols=3, 
                         comparing_sample_lr='../../test_data//Defective_products_on_line_20%//acc_data//000045_lr.xlsx',
                         comparing_sample_ud='../../test_data//Defective_products_on_line_20%//acc_data//000045_ud.xlsx'
                         )
    >>>     acc_processing_raw(dir='../../test_data//20240911_good_samples//acc_data//', analysis_mask=0b0100, fs=1024, nperseg=int(51200/44), noverlap=10,
                         domain='order', nfft=1024, cols=3, average='None', psd_result_filename='psd_window.xlsx')
    >>>     acc_processing_raw(dir= dir + 'acc_data_100%//', analysis_mask=0b0100, fs=1024, nperseg=int(51200/300), noverlap=10,
                         domain='order', nfft=1024, cols=3, average='None', psd_result_filename= dir + 'psd_window.parquet', file_export_func = to_parquet)
    >>>     acc_processing_raw(dir='../../test_data//20240911_good_samples//acc_data_20%//', analysis_mask=0b0100, fs=1024, nperseg=int(51200/44), noverlap=10,
                         domain='order', nfft=10240, cols=3, average='mean', psd_result_filename='../../test_data//20250331_test//psd.xlsx',
                         estimated_frame_len=int(51200/44), resample_len=1024, NumRotations=10)
    >>>     acc_processing_raw(dir='../../test_data//20250613_test_samples//acc_data_100%//', analysis_mask=0b0100, fs=256, nperseg=int(51200/300), noverlap=10,
                         domain='order', nfft=2560, cols=3, average='mean', psd_result_filename='../../test_data//20250613_test_samples//psd_100%_high_resolution.xlsx',
                         estimated_frame_len=int(51200/300), resample_len=256, NumRotations=10)
    can also be used to process the mic data, which has the same format as acc data
    >>>     acc_processing_raw(dir='../../test_data//20250623_test_samples//mic_data_100%//', analysis_mask=0b0100, fs=256, nperseg=int(51200/300), noverlap=10,
                         domain='order', nfft=256, cols=1, average='mean', psd_result_filename='../../test_data//20250623_test_samples//psd_100%.xlsx',
                         estimated_frame_len=int(51200/300), resample_len=256, NumRotations=10, fg_column=1) 
    """
    if analysis_mask & 0b0001:
        df_stats = pd.DataFrame()
    
    comparing_sample_lr_df = read_acc_file(comparing_sample_lr, usecols=usecols) if analysis_mask & 0b1000 else pd.DataFrame()
    comparing_sample_ud_df = read_acc_file(comparing_sample_ud, usecols=usecols) if analysis_mask & 0b1000 else pd.DataFrame()

    total_time_ms = 0
    total_samples = 0
    for file_name in os.listdir(dir):
        if file_name.endswith('.xlsx') or file_name.endswith('parquet') or file_name.endswith('parquet.gzip'):
            total_samples += 1
            t1 = time.time()
            df = read_acc_file(dir+file_name, usecols=usecols)
            print("read %s"%file_name)
            acc_processing_df(df = df, analysis_mask=analysis_mask, sheet_name=file_name[:-5],
                              df_stats=df_stats if analysis_mask & 0b0001 else None, 
                              coherence_compare_df=comparing_sample_lr_df if 'lr' in file_name else comparing_sample_ud_df,
                              fft_result_filename=fft_result_filename, psd_result_filename=psd_result_filename,
                              Cxy_result_filename=Cxy_result_filename, **arg)
            t2 = time.time()
            total_time_ms += (t2 - t1) * 1000  # convert to milliseconds
    if total_samples > 0:
        print(f'Average preprocessing time per sample: {total_time_ms / total_samples:.2f} ms')
    if analysis_mask & 0b0001:
        df_stats.to_excel(state_result_filename, sheet_name='state')

def savefftplot(df_fft:pd.DataFrame, sample:list, annotate_peaks:bool, annotate_bends:bool, save_fig:bool, save_dir:str):
    '''
    show or save a fft plot of selected sample numbers

    Parameters
    -------
    df_fft : each column represents a accelerometer fft spectrum, first six number is the part number, and there are 8 column for each part,
        which is left/right/axile/fg/up/down/axile/fg, fg signal will not be shown in the figure, so the cols = [0,1,2,4,5,6]
    sample : the selected sample number to show or save the picture, array of integer
    annotate_peaks : True to annotate peaks of fft spectrum
    annotate_bends : True to annotate frequency bends
    save_fig : True to save the fig 
    save_dir : save the figure to this directory
    '''
    for n in sample:
        series = df_fft[df_fft.columns[8*n]].to_numpy()
        cols = [8*n,8*n+1,8*n+2,8*n+4,8*n+5,8*n+6]
        ax = df_fft.iloc[:, cols].plot(title="FFT ", xlabel="Frequency (Hz)", ylabel="Amplitude", logy=True, xlim=(0,5000), layout="constrained",figsize=(10,10))
        if annotate_peaks:
            peak_idxs = annotatePeaks(a=series, freq=df_fft.index, ax=ax, prominence=0.25*series)
        if annotate_bends:
            fb0 = bearingFaultBands(rotationSpeed(df_fft, n), 6, 1.5875, 5.645, 0, width=0.1)
            annotateFreqBands(ax, fb0, df_fft.index)
        if save_fig:
            plt.savefig(save_dir+df_fft.columns[8*n][:6], transparent=False, dpi=80, bbox_inches="tight")
        plt.show()

class BearingFaultBands:
    class Info:
        def __init__(self):
            self.Centers = []
            self.Labels = []
            self.FaultGroups = []
        def __str__(self):
            return "Centers: %s,\nLabels: %s,\nFaultGroups: %s"%(self.Centers, self.Labels, self.FaultGroups)
    def __init__(self):
        self.fault_bands = np.ndarray([], dtype=float)
        self.info = BearingFaultBands.Info()
    
    def __str__(self):
        print('fault bands = \n%s'%self.fault_bands)
        print('info = struct with fields: \n%s'%self.info)
        return ''
        
    def insertDict(self, info_insert:list):
        self.info.Centers.append(info_insert[0])
        self.info.Labels.append(info_insert[1])
        self.info.FaultGroups.append(info_insert[2])
    
    def countWidth(self, width: float):
        fault_bands_list = []
        for i in self.info.Centers:
            fault_bands_list.append([i - width, i + width])
        self.fault_bands = np.array(fault_bands_list)
    
    def countDomain(self,fr:float, domain:Literal["frequency", "order"] = "frequency"):
        self.info.Centers = np.array(self.info.Centers)
        if domain == "frequency":
            self.info.Centers *= fr
            self.fault_bands *= fr

def bearingFaultBands(fr:float, nb:int, db:float, dp:float, beta:float, harmonics = [1], sidebands = [0], width:float = 0.1, domain:Literal["frequency", "order"] = "frequency"):
    '''
    python version of matlab `bearingFaultBands()`[1]_. the calculation is based on fixed outer race with rotating inner race.
    
    Parameters
    ----------
    fr : Rotational speed of the shaft or inner race, this parameter is used if the domain is 'frequency'.
    nb : Number of balls or rollers
    db : Diameter of the ball or roller
    dp : Pitch diameter
    beta : Contact angle in degree
    harmonics : harmonics of the fundamental frequency to be included
        1 (default) | vector of positive integers
    Sidebands : Sidebands around the fundamental frequency and its harmonics to be included
        0 (default) | vector of nonnegative integers
    width : width of the frequency bands centered at the nominal fault frequencies
    domain : units of the spectrum x labels
        * 'frequency' : hz
        * 'order' : transfer the angular rotation to order. The horizontal coordinate of the output spectrum is the amplitude 
            versus the multiples of the inner race rotation cycle, fr. 
    
    Returns
    -------
    fb: Fault frequency bands, returned as an N-by-2 array, where N is the number of fault frequencies. 
        FB is returned in the same units as FR, in either hertz or orders depending on the value of 'Domain'. 
        Use the generated fault frequency bands to extract spectral metrics using faultBandMetrics. 
        The generated fault bands are centered at:
        * Outer race defect frequency, Fo, and its harmonics
        * Inner race defect frequency, Fi, its harmonics and sidebands at FR
        * Rolling element (ball) defect frequency, Fbits harmonics and sidebands at Fc
        * Cage (train) defect frequency, Fc and its harmonics
        The value W is the width of the frequency bands, which you can specify using the 'Width' name-value pair.
    Info: Information about the fault frequency bands in FB, returned as a structure with the following fields:
        * Centers — Center fault frequencies
        * Labels — Labels describing each frequency
        * FaultGroups — Fault group numbers identifying related fault frequencies
    
    .. [1] https://wsww.mathworks.com/help/predmaint/ref/bearingfaultbands.html
    '''
    alpha = np.cos(beta * np.pi / 180)
    Fc_order = 0.5 * (1 - db/dp * alpha)
    Fb_order = 0.5 * (dp/db - db/dp * alpha**2)
    Fo_order = nb * Fc_order
    Fi_order = nb * (1 - Fc_order)
    res = BearingFaultBands()
    for i in harmonics:
        res.insertDict([Fo_order * i, '%dFo'%i, 1])
        for j in sidebands:
            fi = Fi_order * i
            fb = Fb_order * i
            if j > 0:
                res.insertDict([fi - j,             '%dFi-%dFr'%(i,j), 2])
            res.insertDict([    fi,                 '%dFi'%i,          2])
            if j > 0:
                res.insertDict([fi + j,             '%dFi+%dFr'%(i,j), 2])
                res.insertDict([fb - j * Fc_order,  '%dFb-%dFc'%(i,j), 3])
            res.insertDict([    fb,                 '%dFb'%i,          3])
            if j > 0:
                res.insertDict([fb + j * Fc_order,  '%dFb+%dFc'%(i,j), 3])
        res.insertDict([Fc_order * i, '%dFc'%i, 4])
    res.countWidth(width)
    res.countDomain(fr, domain)
    return res

def rotationSpeed(df_fft: pd.DataFrame, sample_no:int):
    speed_lr = df_fft.iloc[:,sample_no * 8 + 3].idxmax() * 0.5
    speed_ud = df_fft.iloc[:,sample_no * 8 + 7].idxmax() * 0.5
    return (speed_lr + speed_ud) / 2

def binary_search(arr, low, high, x):
    '''
    given an array, and its lowest index and highest index, return index closest to the value x
    '''
    if high >= low:
        mid = (high + low) // 2
        if arr[mid] == x:
            return mid
        elif arr[mid] > x:
            return binary_search(arr, low, mid - 1, x)
        else:
            return binary_search(arr, mid + 1, high, x)
    else:
        #if x != arr[low]:
        #    print("binary search: find %f and get %f"%(x, arr[low]))
        return high

def annotateFreqBands(axes: matplotlib.axes.Axes, fb: BearingFaultBands, alpha):
    '''
    Annotate bearing frequency bands in different color based on its fault group.
    Note: the input axes should use constrained layout
    
    Parameters
    ----------
    axes : subplots for annotation
    fb : bearing fault bands
    x : x-axis of subplots, which is an array with a specified range and step increment
    '''
    left, right = axes.get_xlim()
    step = (fb.fault_bands[0][1] - fb.fault_bands[0][0])/2
    x = np.arange(left, right, step)
    length = len(x)
    color_arr = ['red', 'orange', 'green', 'blue']
    mask_arr = []
    for i in range(0,4):
        mask_arr.append(np.zeros(length))
    
    for (x1, x2), g, s in zip(fb.fault_bands, fb.info.FaultGroups, fb.info.Labels):
        if (x1 > x[-1]): 
            continue
        mask_arr[g - 1][binary_search(x, 0, length - 1, x1):binary_search(x, 0, length - 1, x2)] = 1
        axes.annotate(s, xy=(x1, 0.95), xycoords=('data', 'subfigure fraction'), rotation='vertical', verticalalignment='top')
    
    for i in range(0, 4):
        axes.fill_between(x, 0, 1, where= mask_arr[i], color= color_arr[i], alpha=alpha, transform=axes.get_xaxis_transform())
    
    axes.annotate('Cage defect frequency, Fc\nBall defect frequency, Fb\nOuter race defect frequency, Fo\nInner race defect frequency, Fi\n',
                  xy=(0.95, 0.05), xycoords='subfigure fraction', horizontalalignment='right')
    
def techometer(fg_signal: pd.DataFrame, thereshold:float, fs:int, pulse_per_round: int):
    '''
    Parameters
    ----------
    fg_signal : square wave signal array
    thereshold : count for rising edge
    fs : smapling frequency
    pulse_per_round : pulse numbers per round, used for rotation speed calculation
    
    returns
    ------
    rotating speed rps= round per second (hz)
    '''
    state = False
    delta = 0
    rps = pd.Series(np.zeros(len(fg_signal)))
    time_buffer = fs/200
    for i in range(1, len(fg_signal), 1):
        delta += 1
        if fg_signal[i] > thereshold and state == False:
            # preventing zero as denominator
            if (i < time_buffer):
                continue
            rps[i] = fs / (delta * pulse_per_round)
            delta = 0
            state = True
        else:
            rps[i] = rps[i - 1]
            if fg_signal[i] < thereshold:
                state = False
    return rps

def cycle_detect(fg_signal:pd.DataFrame, thereshold:float, pulse_per_round:int):
    '''
    Parameters
    ----------
    fg_signal : square wave signal array
    thereshold : count for rising edge
    pulse_per_round : pulse numbers per round
    
    returns
    ------
    index : list(int)
        marked the begining of each cycle
    '''
    counter = -1
    state = (fg_signal[0] > 0)
    cycle_begin_index = []
    for i in range(len(fg_signal)):
        if fg_signal[i] > thereshold:
            if state == False:
                state = True
                counter += 1
                if counter % pulse_per_round == 0:
                    cycle_begin_index.append(i)
                    counter %= pulse_per_round
        else:
            state = False
    return cycle_begin_index

def slice_frame(input:pd.DataFrame, fg_column:int, threshold:float, pulse_per_round:int, estimated_frame_len:int, resample_len:int, NumRotations:int, cols:int):
    '''
    use fg_signal and pulse_per_round to determine the begine of each cycle, returned the resampled and time-synchronous averaged frame.
    
    Parameters
    ----------
    estimated_frame_len : assume the rotation is in the same speed. Remove outliers exceed the speed varying tolerence 20%.
    NumRotations : the output signal frame is averaged every NumRotations
    
    Notes
    -----
    https://www.mathworks.com/help/signal/ug/vibration-analysis-of-rotating-machinery.html
    '''
    cycle_begin_index = cycle_detect(input.iloc[:,fg_column], threshold, pulse_per_round)
    # determine a proper upr_limit and lwr_limit
    upr_limit = int(estimated_frame_len * 1.2)
    lwr_limit = int(estimated_frame_len * 0.8)
    # resample to specified length
    if resample_len is None:
        resample_len = estimated_frame_len
        print('set resample length = %d'%estimated_frame_len)
    cycle_frame = None
    for i in range(1, len(cycle_begin_index)):
        length = cycle_begin_index[i] - cycle_begin_index[i - 1]
        if length > lwr_limit and length < upr_limit:
            original_cycle = input.iloc[cycle_begin_index[i - 1]:cycle_begin_index[i], :cols]
            resampled_cycle = signal.resample(original_cycle, resample_len)
            if cycle_frame is None:
                cycle_frame = resampled_cycle.reshape((1, resampled_cycle.shape[0], resampled_cycle.shape[1]))
            else:
                resampled_cycle = resampled_cycle.reshape((1, resampled_cycle.shape[0], resampled_cycle.shape[1]))
                cycle_frame = np.concatenate((cycle_frame, resampled_cycle), axis=0)
    # time synchronous average for every NumRotations frame
    for i in range(cycle_frame.shape[0] - NumRotations):
        cycle_frame[i, :, :cols] = np.mean(cycle_frame[i:i+NumRotations, :, :cols], axis=0)
    
    return cycle_frame[:-NumRotations, :, :cols]

def fg_fft(fg_signal: pd.DataFrame, fs = 1, nperseq=8192, noverlap=8192//2):
    '''
    use fg signal to get the fft and find the rotation speed in hz
    '''
    detrend_df = fg_signal - np.mean(fg_signal.to_numpy(), axis=0)
    frames = librosa.util.frame(detrend_df, frame_length=nperseq, hop_length=int(nperseq-noverlap))
    window = np.hanning(nperseq)
    windowed_frames = np.empty(frames.shape)
    for col in range(frames.shape[-1]):
        np.multiply(window, frames[:,col], out=windowed_frames[:,col])
    sp = np.fft.rfft(windowed_frames, n=nperseq, norm='backward', axis=0)
    freq = np.fft.rfftfreq(n=nperseq, d=1./fs)
    rps = pd.Series(np.zeros(len(fg_signal)))
    idx = np.argmax(np.abs(sp), axis=0)
    hop_lenth = int(nperseq - noverlap)
    start = 0
    for i in idx:
        rps[start:start + hop_lenth] = freq[i]/2
        start += hop_lenth
    # ending samples
    rps[start:] = rps[start - 1]    
    return rps

def level_and_rpm_seperate_processing(hdf_level_time_filename:str, level_sheet:str, level_col:list, fs = 48000, 
                                      hdf_rpm_time_filename = None, rpm_sheet = None, fs_rpm = 1000,
                                      nperseq = 8192, overlap = 0.75, fft_filename = None, fft_sheet = None):
    '''
    read seperate **level_vs_time.hdf** and **rpm_vs_time.hdf**, with different sampling frequency, to calculate
    FFT based on rpm rotating speed, in order to compare different sample with normalized rotating frequency order
    and output the FFT order result to specified file

    Parameters
    ----------
    hdf_level_time_filename : level_vs_time.hdf transported excel
    level_sheet : sheet name
    level_col : used columns
    fs : sampling frequency
    hdf_rpm_time_filename : rpm_vs_time.hdf transported excel
    rpm_sheet : sheet name
    fs_rpm : rpm_vs_time.hdf sampling frequency, if it is lower than lever_vs_time, than use duplicate for sample augmentation
    nperseq : number of sample per frame
    overlap : percentage of overlape
    fft_filename : output fft file name
    fft_sheet : output fft sheet name, can be append to an exist fft result file as seperate sheet 

    Examples
    --------
    >>> sound_hdf = '../../test_data//20240808//good-100%-18300.Level vs. Time.xlsx'
    >>> rpm_hdf = '../../test_data//20240814//1833-20%.RPM vs. Time.xlsx'
    >>> fft_file = '../../test_data//20240808//fft_order.xlsx'
    >>> level_and_rpm_seperate_processing(hdf_level_time_filename=sound_hdf, hdf_rpm_time_filename=rpm_hdf, level_sheet='Sheet22', 
    ...                                   level_col=[0,1], rpm_sheet='Sheet1',
    ...                                   fft_filename=fft_file, fft_sheet='1833')
    '''
    workbook = openpyxl.load_workbook(hdf_level_time_filename, read_only=True, data_only=True, keep_links=False)
    title = workbook[level_sheet]["B5"].value
    df = pd.read_excel(hdf_level_time_filename, sheet_name=level_sheet, header=0, index_col=0, skiprows=13, usecols=level_col)
    # rewrite column title adding title
    df.rename(columns=lambda x:title.split()[0], inplace=True)
    df_rpm = pd.read_excel(hdf_rpm_time_filename, sheet_name=rpm_sheet, header=0, index_col=0, skiprows=13, usecols="A:B")
    # calculate frame
    df_rpm = df_rpm.loc[df_rpm.index.repeat(fs/fs_rpm)]
    df_rpm.reset_index(drop=True, inplace=True)
    frames = librosa.util.frame(df_rpm, frame_length=nperseq, hop_length=int(nperseq*(1-overlap)), axis=0)
    rps = []
    for frame in frames:
        rps.append(np.round(np.mean(frame, axis=0)/60))
    df_fft = get_fft(df, cut_off_freq = 10, fs = fs, frame_len=nperseq, overlap = overlap,
                     domain="order", pulse_per_round = 2, rps = rps, cols = 1)
    to_excel(df_fft, fft_sheet, fft_filename)
    
def compare_rps_of_rpm_vs_time_file(dir):
    df_dict = {}

    for file_name in os.listdir(dir):
        if file_name.endswith('.xlsx') and not file_name.startswith('~$'):
            print('opening file: %s'%file_name)
            wb = openpyxl.load_workbook(dir + file_name, read_only=True, data_only=True, keep_links=False)
            title = wb['Sheet1']["B5"].value
            wb.close()
            df = pd.read_excel(dir + file_name, sheet_name='Sheet1', header=0, index_col=0, skiprows=13)
            df[df.columns[0]] = df[df.columns[0]]/60
            key = title.split()[0]
            df.rename(columns=lambda x: title.split()[0], inplace=True)
            df_dict[key] = df
    return df_dict

def csd_order(x:pd.DataFrame, y:pd.DataFrame, nperseg:int, noverlap:int, cols:int, nfft = None, average:Literal['mean', 'median', 'None'] = 'mean', **arg):
    '''
    the time domain is transfered to order domain, every data sample in the input signal represnet the same increment of rotation angle,
    x and y input signal has the same sampling angle.
    
    Parameters
    ----------
    nperseg : number of samples of each window frame. In the order domain, this is estimated by fs/average_rotating_frequency.
    noverlap : overlapping samples. In the order domain, this is used for time synchronized average for every noverlap cycles.
    nfft : int, optional
        Length of the FFT used, if a zero padded FFT is desired. If
        `None`, the FFT length is `nperseg`. Defaults to `None`.
        controlled the frequency resolution of the spectrum, using zero-padding to increase the resolution as nfft/nperseg
    cols : use first int(cols) for fft analysis, this values usually equal to the number of accelerometers.
    average : { ‘mean’, ‘median’, 'None' },
                    Method to use when averaging periodograms. If the spectrum is complex, the average is computed separately for the real and imaginary parts. Defaults to ‘mean’.
    '''
    win = np.hanning(nperseg) if nfft == None else np.hanning(nfft)

    # detrend
    signal.detrend(x, axis=0, type='constant', overwrite_data=True) 
    freq_order_x, fft_x = fft(df=x, nperseg=nperseg, window = win, noverlap=noverlap, nfft=nfft, domain='order', cols=cols, **arg)
    if x.equals(y):
        fft_y = fft_x
    else:
        # detrend
        signal.detrend(y, axis=0, type='constant', overwrite_data=True) 
        _, fft_y = fft(df=y, nperseg=nperseg, window = win, noverlap=noverlap, nfft=nfft, domain='order', cols=cols, **arg)    

    frame_len = min(fft_x.shape[0], fft_y.shape[0])
    Pxy = np.empty((frame_len, fft_x.shape[1], cols), dtype=np.complex128)
    
    # scaling for power spectral density
    fs = nperseg if nfft == None else nfft
    scale = 1.0 / (fs * (win*win).sum())
    # input signal is real so the rfft amplitude *2, and if nperseq can not divided by 2, the last point is unpaired Nyquist freq point, don't double
    not_divided_by_2 = nperseg % 2

    Pxy[:,:,:] = np.conjugate(fft_x[:frame_len,:,:cols]) * fft_y[:frame_len,:,:cols] * scale
    if not_divided_by_2:
        Pxy[:,1:-1,:] *= 2
    else:
        Pxy[:,1:,:] *= 2

    # outputting for debug purpose
    #fft_frame_to_excel(Pxy, sheet_names=['Pxy'], fft_filename='Pxy.xlsx', index=freq_x)
    
    # Average over windows
    bias = signal._spectral_py._median_bias(fft_x.shape[0])

    if average == 'median':
        if np.iscomplexobj(Pxy):
            Pxy = (np.median(np.real(Pxy), axis=0) + 1j * np.median(np.imag(Pxy), axis=0))
        else:
            Pxy = np.median(Pxy, axis = 0)
        Pxy /= bias
    elif average == 'mean':
        Pxy = np.mean(Pxy, axis=0)
    elif average == 'None':
        print('return Pxy without averaging over %d windows'%Pxy.shape[0])
    else:
        raise ValueError('choose from specified methods')
    
    return freq_order_x, Pxy

def coherence(x:pd.DataFrame, y:pd.DataFrame, fs:int, nperseg:int, noverlap:int, cols:int, domain:Literal['frequency', 'order'] = 'order',
              nfft=None, average:Literal['mean', 'median'] = 'mean', visualize=False):
    '''
    rewrite signal.coherence adding domain parameter to get frequency as order of rotating frequency
    adding the average parameter for csd calculation 

    Parameters
    ----------
    x : array_like. Time series of measurement values
    y : array_like. Time series of measurement values
    fs : float, optional. Sampling frequency of the x and y time series.
    nperseg : int. Length of each segment. length of the window.
    noverlap : int. Number of points to overlap between segments.
    cols : use first int(cols) for fft analysis, this values usually equal to the number of accelerometers.
    domain : units of the spectrum x labels
                'frequency': hz
                'order': transfer the angular rotation to order. The horizontal coordinate of the output spectrum is the amplitude 
                         versus the multiples of the inner race rotation cycle, fr. 
    fg_column : fg signal as square wave, usually the last column number of input data frame, unused if rps is given.
    average : { ‘mean’, ‘median’ },
                    Method to use when averaging periodograms. If the spectrum is complex, the average is computed separately for the real and imaginary parts. Defaults to ‘mean’.
    visualize : Set True to show the coherence graph
    
    Notes
    -----
    refence of coherence:
    1. https://atmos.washington.edu/~dennis/552_Notes_ftp.html Cross Spectrum Analysis Section 6c
    2. https://ocw.mit.edu/courses/6-011-introduction-to-communication-control-and-signal-processing-spring-2010/pages/readings/ Chapter 9,10,11
    3. https://www.nii.ac.jp/qis/first-quantum/forStudents/lecture/pdf/noise/chapter1.pdf
    '''
    
    if domain == 'frequency':
        # same as signal.coherence but add param to change averaging method
        freqs, Pxx = signal.welch(x.iloc[:,:cols], fs=fs, window='hann', nperseg=nperseg, noverlap=noverlap, axis=0, average=average)
        _, Pyy = signal.welch(y.iloc[:,:cols], fs=fs, window='hann', nperseg=nperseg, noverlap=noverlap, axis=0, average=average)
        _, Pxy = signal.csd(x.iloc[:,:cols], y.iloc[:,:cols], fs=fs, window='hann', nperseg=nperseg, noverlap=noverlap, axis=0, average=average)
        # organize into dataframe
        Cxy = pd.DataFrame(data=np.abs(Pxy)**2 / Pxx / Pyy, index=freqs)
        Cxy.index.rename('frequency [Hz]', inplace=True)
        x_limit = 5000
    else:
        # order of rotating frequency
        orders_x, Pxx = csd_order(x=x, y=x, fs=fs, nperseg=nperseg, noverlap=noverlap, cols=cols, nfft=nfft, average=average)
        _, Pyy = csd_order(x=y, y=y, fs=fs, nperseg=nperseg, noverlap=noverlap, cols=cols, nfft=nfft, average=average)
        _, Pxy = csd_order(x=x, y=y, fs=fs, nperseg=nperseg, noverlap=noverlap, cols=cols, nfft=nfft, average=average)
        
        Cxy = pd.DataFrame(np.abs(Pxy)**2 / np.real(Pxx) / np.real(Pyy), index=orders_x)
        Cxy.index.rename('order of rotating frequency', inplace=True)
        x_limit = orders_x[-1]
    
    column_name = ['%s vs %s'%(x.columns[i], y.columns[i]) for i in range(cols)]
    Cxy.columns = column_name

    # visualize
    if visualize:
        fig, axs = plt.subplots(cols, 1, layout='constrained', sharex=True)
    
        for i in range(cols):
            ax = axs if cols == 1 else axs[i]
            Cxy.iloc[:,i].plot(ax=ax, legend=True, xlabel=Cxy.index.name, ylabel='Coherence', logx=False, logy=False, xlim=(0,x_limit))
            ax.grid(visible=True, which='both', axis='both')
            if i == 0:
                ax.set_title('Coherence (average type: %s)'%average)
        plt.show()
    
    return Cxy

def corr(df:pd.DataFrame, result_filename:str):
    for meth in ['pearson', 'kendall', 'spearman']:
        df_corr = df.corr(method=meth)
        to_excel(df_corr, meth, result_filename)

def fft_frame_to_excel(fft_frame:np.ndarray, sheet_names:list, fft_filename:str, index:np.ndarray):
    '''
    output a not averaged fft frames into excel, the shape of input frames should be [windowed_frame, frequency, column]
    the column >= 1 represnets the input channels

    Parameters
    ----------
    sheet_names : if the fft_frame has multiple channels, save each channels in seperate sheet
    '''
    for sheet in range(len(sheet_names)):
        df_fft = pd.DataFrame(fft_frame[:,:,sheet].transpose(), index=index, columns=range(fft_frame.shape[0]))
        to_excel(df=df_fft, sheet_name=sheet_names[sheet], filename=fft_filename)

def coherence_test(average:Literal['mean', 'median'] = 'mean', plot_mask=0b10000):
    np.random.seed(0)
    fs = 51200
    frame_len = 8192
    #f1 = 40
    #f2 = 50
    #x = pd.DataFrame(10 * np.sin(2 * np.pi * f1 * t) + 10 * np.sin(2 * np.pi * 2 * f1 * t) + 10 * np.sin(2 * np.pi * 3 * f1 * t) + 0.5 * np.random.randn(n), columns=['x'])
    #y = pd.DataFrame(15 * np.sin(2 * np.pi * f2 * t + np.pi / 4) + 10 * np.sin(2 * np.pi * 2 * f2 * t) + 10 * np.sin(2 * np.pi * 3 * f2 * t) + 0.5 * np.random.randn(n), columns=['y'])
    x = pd.read_excel('../../test_data//Defective_products_on_line_20%//acc_data//000045_lr.xlsx', header=0, usecols="A:D")
    y = pd.read_excel('../../test_data//20240911_good_samples//acc_data//000022_lr.xlsx', header=0, usecols="A:D")
    
    # visualize
    mask = plot_mask
    count = 0
    while mask>0:
        count += 1 if mask & 1 else 0
        mask = mask >> 1
    fig, axs = plt.subplots(count, 1, layout='constrained')    
    idx = 0
    axs = [axs] if (count == 1) else axs # change the axs into list
    if plot_mask & 0b00001:
        # input
        axs[idx].plot(np.arange(x.shape[0])/fs, x, label=x.columns)
        axs[idx].plot(np.arange(y.shape[0])/fs, y, label=y.columns)
        axs[idx].set_xlabel('time(s)')
        axs[idx].set_ylabel('input signal')
        axs[idx].set_xlim(0, 0.25)
        axs[idx].legend()
        idx += 1
    if plot_mask & 0b00010:
        # fft
        fft_x = get_fft(df=x, fs=fs, nperseg=frame_len, noverlap=0.75*frame_len)
        fft_y = get_fft(df=y, fs=fs, nperseg=frame_len, noverlap=0.75*frame_len)
        axs[idx].plot(fft_x.index, fft_x, label=x.columns)
        axs[idx].plot(fft_y.index, fft_y, label=y.columns)
        axs[idx].legend()
        axs[idx].set_xlim(0, 5000)
        axs[idx].set_yscale('log')
        axs[idx].set_title('input signal fft spectrum at first frame')
        axs[idx].set_xlabel('frequency (Hz)')
        annotatePeaks(x=fft_x.index.to_numpy(), y=fft_x.iloc[:,0].to_numpy(), ax= axs[idx], prominence=fft_x.iloc[:, 0].to_numpy()*0.9,
                      rotation = 0, dot = 'x', xytext=(0, 0), arrowprops=None)
        annotatePeaks(x=fft_y.index.to_numpy(), y=fft_y.iloc[:,0].to_numpy(), ax= axs[idx], prominence=fft_y.iloc[:, 0].to_numpy()*0.9,
                      rotation = 0, dot = 'x', xytext=(0, 0), arrowprops=None)
        idx += 1
    if plot_mask & 0b00100:
        # csd
        freq_csd2, Pxy2 = signal.csd(x, y, fs, nperseg=frame_len, noverlap=frame_len*0.75, axis=0, average=average)
        axs[idx].plot(freq_csd2, np.abs(Pxy2[:, :3]), label = ['%s vs %s'%(x.columns[i], y.columns[i]) for i in range(3)])
        axs[idx].set_xlim(0, 500)
        axs[idx].set_yscale('log')
        axs[idx].set_xlabel('frequency (Hz)')
        axs[idx].set_ylabel('CSD [V**2/Hz]')
        axs[idx].legend()
        annotatePeaks(x=freq_csd2, y=np.abs(Pxy2[:, 0]), ax= axs[idx], prominence=np.abs(Pxy2[:,0])*0.9, rotation = 0,
                      annotateY=True, dot = 'x', xytext=(0, 0), arrowprops=None)
        idx += 1
    if plot_mask & 0b01000:
        # csd_order
        freq_csd, Pxy = csd_order(x=x, y=y, fs=1024, nperseg=1024, noverlap=10, cols=3)
        axs[idx].plot(freq_csd, np.abs(Pxy), label = ['%s vs %s'%(x.columns[i], y.columns[i]) for i in range(3)])
        axs[idx].set_xlim(0, Pxy.shape[0])
        axs[idx].set_yscale('log')
        axs[idx].set_xlabel('order')
        axs[idx].set_ylabel('CSD [V**2/Hz]')
        axs[idx].legend()
        annotatePeaks(x=freq_csd, y=np.abs(Pxy[:,0]), ax= axs[idx], prominence=np.abs(Pxy[:,0])*0.9, rotation = 0,
                      annotateY=True, dot = 'x', xytext=(0, 0), arrowprops=None)
        idx += 1
    if plot_mask & 0b10000:
        # coherence
        Cxy = coherence(x=x, y=y, fs=1024, nperseg=1163, noverlap=10, cols=3, domain = 'order', average=average, nfft=1024)
        Cxy.plot(ax=axs[idx], logy=True, xlabel='order', ylabel='Coherence', xlim=(0, Cxy.shape[0]))
    plt.show()

def class_label(sample_num:str):
    '''
    return the classification type
    * 0: normal
    * 1: bearing_noise
    * 2: unkwown_noise
    '''
    normal = ['000022','000027','000030','000037','000039','000045','000048','000050','000051','000052','000053']
    bearing_noise = ['003720','003735','003861','004072','004073','004802']
    unknown_noise = ['000785','001833','002577','004124']
    if sample_num in normal:
        return 0
    elif sample_num in bearing_noise:
        return 1
    elif sample_num in unknown_noise:
        return 2
    else:
        raise ValueError('undefined sample number')

def compare_spectrum_plot(df:pd.DataFrame, high_light:bool = False, titles:list = ['left','right','lr_axial', 'up', 'down','ud_axial'],
                          xmax = None, mode:Literal['coherence', 'normal'] = 'normal', label_method = class_label, **arg):
    '''
    read exported excel file of spectrum, plot with seperated sensor channel, and the color distinguished with abnormal type

    Parameters
    ----------
    df : spectrum data, can be fft, psd, coherence...
    high_light : whether to high light the frequency where all normal samples are greater or lower than abnormal samples
    titles : controls which sensor position data shows up
    xmax : max value of the x axis, default None.
    mode : 
        * coherence : y axis use **linear** scale and high light all nonoverlapping groups
        * normal : y axis use **log** scale and high light all nonoverlapping groups
    **arg : parameters pass to plot()

    Example
    -------
    >>> compare_spectrum_plot(df = read_sheets('coherence.xlsx'))
    >>> compare_spectrum_plot(df = read_sheets('../../test_data//Defective_products_on_line_20%//fft_abnormal.xlsx'), high_light=True, linewidth=1, alpha=0.5)
    >>> compare_spectrum_plot(df = read_sheets('../../test_data//20240911_good_samples//fft.xlsx'), high_light=True, linewidth=1, alpha=0.5)
    >>> compare_spectrum_plot(df = read_sheets('psd.xlsx', usecols=[0,1,2,3], combine=True), high_light=True, linewidth=1, alpha=0.5, xmax = 20)
    used for psd with only one channel
    >>> compare_spectrum_plot(df=read_sheets('../../test_data//20250623_test_samples//psd_20%.xlsx', usecols=[0,1]),
                              titles=['mic'], high_light=True, label_method=label_test)
    '''
    if xmax is not None:
        df = df.iloc[:xmax]
    # difine sample classification
    fig, axs = plt.subplots(len(titles),1, sharex='col', layout='constrained')
    colors = [color['green'], color['orange'], color['blue']]

    if len(titles) == 1:
        axs = [axs] # change the axs into list if only one title is given
    for i in range(len(titles)):
        if mode == 'normal':
            axs[i].set_yscale('log')
        axs[i].set_xlim(0, df.index[-1])
        axs[i].set_title(titles[i])
    
    for j in range(len(df.columns)):
        label_sample = label_method(df.columns[j].split(' ')[-1].split('_')[0])
        for i in range(len(titles)):
            if titles[i] in df.columns[j]:
                axs[i].plot(df.index, df.iloc[:,j], color=colors[label_sample], **arg)
                break
    
    if high_light:
        normal_column = [[] for i in titles]
        abnormal_column = [[] for i in titles]
        
        for j in range(len(df.columns)):
            if label_method(df.columns[j].split(' ')[-1].split('_')[0]) == 0:
                for i in range(len(titles)):
                    if titles[i] in df.columns[j]:
                        normal_column[i].append(j)
                        break
            else:
                for i in range(len(titles)):
                    if titles[i] in df.columns[j]:
                        abnormal_column[i].append(j)
                        break

        bool_high_light = np.zeros((len(df.index), len(titles)), dtype=bool)
        for idx in range(len(df.index)):
            if mode == 'coherence':
                for i in range(len(titles)):
                    if max(df.iloc[idx, abnormal_column[i]]) < min(df.iloc[idx, normal_column[i]]):
                        bool_high_light[idx, i] = True
                        print('%s find at %f'%(titles[i],df.index[idx]))
            else:
                for i in range(len(titles)):
                    if min(df.iloc[idx, abnormal_column[i]]) > max(df.iloc[idx, normal_column[i]]):
                        bool_high_light[idx, i] = True
                        print('%s find at %f'%(titles[i],df.index[idx]))
             
        for i in range(len(titles)):
            axs[i].fill_between(df.index, 0, 1, where= bool_high_light[:,i], color= 'red', alpha=0.5, transform=axs[i].get_xaxis_transform())

    plt.show()

def cosine_similarity(df:pd.DataFrame, sheet_name:str, file_name:str):
    distvec = scipy.spatial.distance.pdist(df.transpose(), metric='cosine')
    m = scipy.spatial.distance.squareform(distvec)
    matrix = pd.DataFrame(1-m, index=df.columns, columns=df.columns)
    to_excel(matrix, sheet_name, file_name)

def boxplot(file_name:str, titles:list = ['left','right','lr_axial', 'up', 'down','ud_axial']):
    '''
    shows boxplot of different catagorical type and channel, to see the distrobution of data
    
    Parameters
    ----------
    file_name : input spectrum
    
    Examples
    --------
    >>> boxplot('psd.xlsx')
    '''
    df = read_sheets(file_name, usecols=[0,1,2,3])
    df.index.rename('Order', inplace=True)
    df = df.transpose()
    # add channel and label column, other column are features, seaborn reads dataframe better
    for name in df.index:
        for title in titles:
            if title in name:
                df.loc[name, 'channel'] = title
                break
        df.loc[name, 'label'] = class_label(sample_num=name.split(' ')[-1].split('_')[0])
    
    # draw boxplot
    feature_per_figure = 10
    stop_order = 150
    
    for start_col in range(0, stop_order, feature_per_figure):
        fig, axs = plt.subplots(len(titles), feature_per_figure, layout='constrained')
    
        for i in range(len(titles)):
            axs[i][0].set_ylabel(titles[i])
            for j in range(feature_per_figure):
                legend = 'auto' if i == 0 and j == 0 else False
                sns.boxplot(x=df.loc[df['channel'] == titles[i], j + start_col], data=df.loc[df['channel'] == titles[i]], hue='label', ax=axs[i][j], legend=legend)
                sns.stripplot(x=df.loc[df['channel'] == titles[i], j + start_col], data=df.loc[df['channel'] == titles[i]], alpha=0.5, hue='label', ax=axs[i][j], legend=legend)
        plt.show()

def calculate_spectral_stats(low_order:int, high_order:int, df:pd.DataFrame):
    '''
    calculate the averaged energy and standard deviation of the specified range of spectrum (from low_idx to high_idx order)
    '''
    idx = len(df.columns) - 1
    while isinstance(df.columns[idx], str):
        idx -= 1
        
    low_idx = binary_search(df.columns, 0, idx, low_order)
    high_idx = binary_search(df.columns, 0, idx, high_order)
    print('calculate between %i and %i'%(low_idx, high_idx))
    df['mean_energy_%i_%i'%(low_order, high_order)] = df.iloc[:, low_idx:high_idx].mean(axis=1)
    df['std_energy_%i_%i'%(low_order, high_order)] = df.iloc[:, low_idx:high_idx].std(axis=1)
    return df 

def pa_to_dB_SPL(pa:pd.DataFrame, reference:float = 20e-6):
    '''
    transfer the pressure level in pascal to dB SPL, using the reference of 20uPa
    '''
    if isinstance(pa, pd.Series):
        pa = pa.to_frame()
    
    # calculate the RMS dB SPL
    rms_pressure_pa = pa.std(axis=0)
    df_dB_SPL = 20 * np.log10(rms_pressure_pa / reference)
    return df_dB_SPL

if __name__ == '__main__':
    pass